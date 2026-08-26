if 'RUN_CONTEXT' in globals():  # Running on Peliqan
    RUN_ENV = 'peliqan'
else:  # Running outside of Peliqan
    RUN_ENV = 'local'
    from peliqan import Peliqan
    import streamlit as st
    import os
    api_key = os.getenv("PELIQAN_API_KEY")
    if not api_key:
        st.error("PELIQAN_API_KEY environment variable is not set.")
        st.stop()
    interface_id = os.getenv("PELIQAN_INTERFACE_ID", 0)
    pq = Peliqan(api_key)
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        if get_script_run_ctx() is not None:
            RUN_CONTEXT = "interactive"
    except Exception:
        RUN_CONTEXT = "background"

"""
ts_monthly_view - keeps ts_reporting.v_monthly_entries current.

Not published. It exists so the monthly export has one agreed definition
of "an entry", and so that definition is refreshed on a schedule instead
of only when somebody happens to click Export.

What it does: one view per calendar month -
ts_reporting.v_monthly_entries_2026_08 and so on - each holding that
month's timetable rows joined to user, task, project and client, ordered
by employee then date. The month window is baked into each view's SQL, so
reading one needs no date filter.

Schedule: monthly, via the schedule configured on this Data App in
Peliqan. Each run asserts the view for the current month AND the one
before it: run on the 1st, the new month's view comes into being and the
month that just closed gets a final pass. Running more often is harmless
- an existing view is updated in place, never duplicated.

Older months are not touched. Their views already exist from earlier runs,
and ts_my_week creates one on demand for any month somebody exports.

The view reads the live ts_prod tables, NOT ts_reporting.fact_timetable.
fact_timetable is a materialized query table and lags writes to
ts_prod.timetable until its own query is re-run - fine for the BI
dashboard, wrong for an export somebody is about to send to a client.
Because this is a plain view, it needs no refresh to be current: the
scheduled run only matters when the SQL itself changes.

This is the only app that exports. ts_my_week (12011) used to carry a
byte-identical copy of the block below and call ensure_monthly_view()
itself; that copy has been removed, so there is nothing left to keep in
sync - edit the block here and nowhere else.
"""

import io
from datetime import date

import pandas as pd

# =====================================================
# Config
# =====================================================

DW_NAME = "dw_3202"
S = "ts_prod"

INTERACTIVE = globals().get("RUN_CONTEXT") == "interactive"


def say(message):
    """Log to the run output, and to the page when opened interactively."""
    print(message)
    if INTERACTIVE and "st" in globals():
        st.write(message)


# =====================================================
# Monthly export view  (SHARED BLOCK - keep byte-identical with
# ts_monthly_view / 12507_ts_monthly_view.py)
#
# Two Data Apps cannot import each other, so this block is duplicated.
# Edit it in one place and copy it to the other.
#
# One view per calendar month: ts_reporting.v_monthly_entries_2026_08 and
# so on. The month window is baked into each view's SQL, so reading one
# needs no date filter - ask for the month by name.
#
# Sourced from the live ts_prod tables on purpose, NOT from
# ts_reporting.fact_timetable: that one is a materialized query table and
# does not reflect writes until its underlying query is re-run. A plain
# view over ts_prod is always current, which is what an export needs.
# =====================================================

VIEW_SCHEMA = "ts_reporting"
VIEW_PREFIX = "v_monthly_entries_"

EXPORT_COLUMNS = [
    ("employee", "Employee"),
    ("entry_day", "Date"),
    ("start_time", "Start"),
    ("duration_hours", "Hours"),
    ("billable", "Billable"),
    ("client", "Client"),
    ("project", "Project"),
    ("task", "Task"),
    ("note", "Note"),
    ("approved", "Validated"),
]


def month_bounds(month_start):
    """(first day, first day of next month) - one view's window."""
    if month_start.month == 12:
        return month_start, date(month_start.year + 1, 1, 1)
    return month_start, date(month_start.year, month_start.month + 1, 1)


def view_name(month_start):
    return f"{VIEW_PREFIX}{month_start.strftime('%Y_%m')}"


def monthly_view_sql(month_start):
    start, end = month_bounds(month_start)
    return f"""
SELECT
    t.id                                      AS entry_id,
    t.user_id::text                           AS user_id,
    COALESCE(u.name, u.email, 'Unknown user') AS employee,
    t.date                                    AS entry_date,
    CAST(t.date AS DATE)                      AS entry_day,
    COALESCE(t.duration, 0)                   AS duration_minutes,
    ROUND(COALESCE(t.duration, 0) / 60.0, 2)  AS duration_hours,
    COALESCE(tk.billable, FALSE)              AS billable,
    COALESCE(c.name, '-')                     AS client,
    COALESCE(p.name, '-')                     AS project,
    COALESCE(tk.name, '-')                    AS task,
    COALESCE(t.internal_description, '')      AS note,
    COALESCE(t.approved, FALSE)               AS approved
FROM {S}.timetable t
LEFT JOIN {S}.users    u  ON u.id::text = t.user_id::text
LEFT JOIN {S}.tasks    tk ON tk.id = t.task_id
LEFT JOIN {S}.projects p  ON p.id = tk.project_id
LEFT JOIN {S}.clients  c  ON c.id = p.client_id
WHERE t.date >= TIMESTAMP '{start.isoformat()} 00:00:00'
  AND t.date <  TIMESTAMP '{end.isoformat()} 00:00:00'
ORDER BY employee, t.date
"""


def ensure_monthly_view(month_start):
    """
    Create this month's view, or update it in place when the SQL changed.
    Idempotent: upsert_query looks the table up by name first and only
    falls back to creating it when it does not exist yet.
    """
    return pq.upsert_query(
        name=view_name(month_start),
        query=monthly_view_sql(month_start),
        schema_name=VIEW_SCHEMA,
        database_name=DW_NAME,
        as_view=True,
    )


def fetch_month(month_start, user_id=None):
    """
    One month's view, ordered employee then date. The view is already
    scoped to the month, so only the employee filter is applied here.
    user_id=None means every employee.
    """
    where = "" if user_id is None else f"WHERE user_id = '{int(user_id)}'"
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT * FROM {VIEW_SCHEMA}.{view_name(month_start)}
        {where}
        ORDER BY employee, entry_date
    """) or []
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    dt = pd.to_datetime(df["entry_date"], errors="coerce")
    df["start_time"] = dt.dt.strftime("%H:%M")
    df["entry_day"] = dt.dt.strftime("%Y-%m-%d")
    return df


def build_workbook(df, sheet_title):
    """The month as .xlsx bytes. openpyxl is available in the runtime."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]                  # Excel's hard limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="053763")
    for col, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(df.to_dict("records") if not df.empty else [], start=2):
        for col, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = row.get(key)
            if key in ("approved", "billable"):
                value = "yes" if value in (True, "true", "True", 1, "1") else "no"
            elif key == "duration_hours":
                value = float(value or 0)
            ws.cell(row=r, column=col, value=value)

    widths = {"employee": 22, "entry_day": 12, "start_time": 8, "duration_hours": 8,
              "billable": 9, "client": 20, "project": 24, "task": 28, "note": 50,
              "approved": 10}
    for col, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 16)

    ws.freeze_panes = "A2"
    if not df.empty:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{len(df) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(month_start, who):
    return f"timesheet_{month_start.strftime('%Y-%m')}_{who}.xlsx"

# =====================================================
# Run
# =====================================================

def previous_month(month_start):
    if month_start.month == 1:
        return date(month_start.year - 1, 12, 1)
    return date(month_start.year, month_start.month - 1, 1)


today = date.today()
this_month = date(today.year, today.month, 1)
targets = [previous_month(this_month), this_month]

dbconn = pq.dbconnect(DW_NAME)
results = []

for month in targets:
    name = f"{VIEW_SCHEMA}.{view_name(month)}"
    try:
        results.append(ensure_monthly_view(month))
    except Exception as exc:
        # One bad month must not stop the other from being refreshed.
        say(f"FAILED {name}: {exc}")
        continue
    try:
        counted = dbconn.fetch(DW_NAME, query=f"""
            SELECT COUNT(*) AS entries, COUNT(DISTINCT employee) AS employees
            FROM {name}
        """) or []
        summary = (f"{counted[0].get('entries')} entries across "
                   f"{counted[0].get('employees')} employees"
                   if counted else "row count unavailable")
    except Exception as exc:
        # A failed count says nothing about whether the view itself is fine.
        summary = f"created/updated, but counting its rows failed: {exc}"
    say(f"{name}: {summary}")

if INTERACTIVE and "st" in globals():
    st.caption("This app is not published - it maintains the monthly views "
               "that ts_my_week's Excel export reads.")
    st.write(results)
