if 'RUN_CONTEXT' not in globals():  # Running outside of Peliqan
    import os
    import streamlit as st
    from peliqan import Peliqan
    api_key = os.getenv("PELIQAN_API_KEY")
    if not api_key:
        st.error("PELIQAN_API_KEY environment variable is not set.")
        st.stop()
    pq = Peliqan(api_key)

"""
ts_my_week (v4.2) - POC

Google-Calendar-style week view for timesheet entries, reload-free.

Access: the Google login is mandatory. Nothing renders but a sign-in
page until the visitor completes a Google login whose verified email
matches a row in ts_prod.users - the same rule ts_mcp_server applies to
its Bearer tokens. Needs three Secret Store entries and the redirect URI
registered in Google Cloud; the login then lives in an encrypted cookie
for 12 hours. The logged-in user IS the viewer.

Monthly Excel export lives in its own Data App (ts_monthly_view), not
here.

Who-sees-what (the same scope model ts_mcp_server enforces, keyed on
ts_prod.users.scope, cumulative):
  - Scope 1 (employee): sees only their OWN weekly calendar. Can add/edit/
    delete entries while the week is open, and Submit / Unsubmit it.
  - Scope 2+ (manager/admin): gets a "Viewing" selectbox to open anyone's
    calendar, read-only, plus Validate on a submitted week.
      - viewing their own week: Submit / Unsubmit, plus Validate once
        the week is submitted (managers may validate their own week).
      - viewing someone else's week: read-only; if that week is
        submitted, a Validate button appears.
  - Validate is the terminal step, defined by validate_week() below:
    every timetable entry of that user in that week is stamped
    approved=true + approved_by, and the submission row becomes
    status='confirmed' with confirmed_by / confirmed_at. Validated entries
    are then locked for everyone, here and in ts_mcp_server, which refuses
    to write to them. (Those columns keep their database names; everything
    a user reads says validated.)

Navigation: a date picker ("Show week of") jumps to the week containing
any date, with explicit "Previous week" / "Next week" buttons next to it
and a "This week" shortcut that only appears when you are looking at
another week.

Adding entries: there is no global New button. Each day column has its own
"+" button in the totals row, and clicking any FREE hour cell in the grid
opens the new-entry dialog prefilled with that day and hour. Hour cells
that are already covered by an entry are not clickable as "add" targets -
otherwise a click on a block could open the add dialog underneath it and
create an overlapping entry.

Interaction model (why Plotly): every click on the grid travels over
Streamlit's websocket as a chart selection - no page reloads, no white
flash. The Peliqan runtime ships Altair 4.2 (predates selection_point),
but Plotly 6.9 + Streamlit 1.49 support on_select, so the grid is a
Plotly figure styled to look like the Altair/Vega original.

Data:
  - ts_prod.users                  read-only (selector + scope)
  - ts_prod.timetable              read + write (viewed user's rows)
  - ts_prod.tasks/projects/clients read-only lookups (labels + colors)
  - ts_prod.timetable_submissions  read + write (submit/validate workflow)

NOTE: st.set_page_config must stay a literal string - the Peliqan runtime
lifts that call into a system prepend that runs before this script body.
"""

import base64
import hashlib
import hmac
import inspect
import io
import json
import secrets
import urllib.error
import urllib.parse
import urllib.request

import plotly.graph_objects as go
import pandas as pd
import streamlit as st
from datetime import date, datetime, time, timedelta, timezone

# NOTE: literal only - executed by the Peliqan runtime BEFORE the script body.
st.set_page_config(page_title="My week", layout="wide")

# Wide layout reserves ~6rem above the first element. This is a calendar:
# the grid wants that vertical space more than the top of the page does.
st.markdown("<style>.block-container{padding-top:2.2rem;}</style>",
            unsafe_allow_html=True)

# =====================================================
# Config
# =====================================================

DW_NAME = "dw_3202"
S = "ts_prod"

MANAGE_SCOPE = 2             # scope >= 2 can view anyone + validate
FIRST_EXPORT_MONTH = date(2026, 7, 1)   # no usable timesheet data before this

# Task ids the viewer may log against, or None for "no restriction".
# Reassigned once the viewer is known, below; the default keeps
# task_selectbox working no matter when a dialog first runs.
TASK_CHOICES = None

DAY_TARGET_MIN = 8 * 60      # minimum per workday
WORKDAYS = (0, 1, 2, 3, 4)
DAY_ABBREV = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# ts_mcp_server's TASK_STATUS, kept in step: a task created here must be
# one the MCP would also accept.
TASK_STATUSES = ["todo", "in_progress", "done", "blocked"]
NEW_TASK_STATUS = "todo"
NEW_TASK_BILLABLE = False

# Project ids the viewer may create a task under, or None for "no
# restriction". Set alongside TASK_CHOICES once the viewer is known.
PROJECT_CHOICES = None

SLOT_MIN = 15                # add-target granularity, in minutes
SLOTS_PER_H = 60 // SLOT_MIN

# Only reached by start_time_input's fallback - see there. st.time_input's
# `step` decides which times are VALID, not merely which ones the menu lists,
# so this is the coarsest step that still lets 13:42 be entered at all.
START_STEP = timedelta(minutes=1)

# Whether st.selectbox can take a value outside its options. Peliqan pins its
# own Streamlit and this app cannot choose it, so ask rather than assume: on a
# runtime without it, passing the argument is a TypeError that takes the whole
# page down. Evaluated once, at import.
CAN_TYPE_NEW_OPTIONS = "accept_new_options" in inspect.signature(st.selectbox).parameters

GRID_START_H = 7             # default visible range; auto-extends to fit
GRID_END_H = 19
ROW_PX = 52                  # pixels per hour row

CLIENT_COLORS = ["#4c78a8", "#54a24b", "#b279a2", "#f58518",
                 "#e45756", "#eeca3b", "#9d755d", "#72b7b2"]

# Stand-in for an entry whose task_id is no longer in the lookup.
LOOKUP_DEFAULT = {"task": "?", "project": "-", "client": "-", "color": "#9d9da6"}

# =====================================================
# Helpers
# =====================================================

def to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def is_true(v):
    return v in (True, "true", "True", 1, "1")


def fmt_dur(minutes):
    minutes = int(minutes or 0)
    h, m = divmod(minutes, 60)
    if h and m:
        return f"{h}h{m:02d}"
    if h:
        return f"{h}h"
    return f"{m}m"


def monday_of(d):
    return d - timedelta(days=d.weekday())


def submission_key(user_id, week_start):
    return f"{int(user_id)}_{week_start.isoformat()}"


def user_display_name(user):
    return (user or {}).get("name") or (user or {}).get("email") or f"user #{(user or {}).get('id')}"

# =====================================================
# Data loading (scoped SQL, cached)
# =====================================================

@st.cache_data(ttl=300)
def load_users():
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, S, "users") or []
    users = [r for r in rows if to_int(r.get("id")) is not None]
    return sorted(users, key=lambda u: str(user_display_name(u)).lower())


@st.cache_data(ttl=60)
def load_entries(user_id):
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT id, task_id, date, duration, internal_description, approved
        FROM {S}.timetable
        WHERE user_id::text = '{int(user_id)}'
    """) or []
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["id", "task_id", "date", "duration",
                                     "internal_description", "approved"])
    df["dt"] = pd.to_datetime(df["date"], errors="coerce")
    df["duration"] = pd.to_numeric(df["duration"], errors="coerce").fillna(0).astype(int)
    return df.dropna(subset=["dt"]).sort_values("dt")


@st.cache_data(ttl=300)
def load_task_lookup():
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT t.id AS task_id, t.name AS task,
               t.project_id, t.status, t.billable, t.description,
               COALESCE(p.name, '-') AS project,
               COALESCE(c.id, 0) AS client_id,
               COALESCE(c.name, '-') AS client
        FROM {S}.tasks t
        LEFT JOIN {S}.projects p ON p.id = t.project_id
        LEFT JOIN {S}.clients  c ON c.id = p.client_id
    """) or []
    lookup = {}
    for r in rows:
        tid = to_int(r.get("task_id"))
        if tid is None:
            continue
        cid = to_int(r.get("client_id")) or 0
        lookup[tid] = {
            "task": r.get("task") or f"Task {tid}",
            "project": r.get("project") or "-",
            "project_id": to_int(r.get("project_id")),
            "status": r.get("status") or "-",
            "billable": is_true(r.get("billable")),
            "description": r.get("description") or "",
            "client": r.get("client") or "-",
            "client_id": cid,
            "color": CLIENT_COLORS[cid % len(CLIENT_COLORS)],
        }
    return lookup


# Deliberately unfiltered above: every task is needed to LABEL an entry,
# including one logged against a client this viewer is not assigned to.
# What a viewer may log against is a separate, narrower set - see
# allowed_task_ids() and TASK_CHOICES.


@st.cache_data(ttl=300)
def load_client_links():
    """
    {client_id: {user_id, ...}} - who is assigned to which client.

    clients.user_list is a many-to-many field and never comes through
    dbconn.fetch(schema, table), so the only way to read it is the
    internal junction table Peliqan itself queries for the relation.
    Same read ts_mcp_server does in _fetch_client_user_links; keep the two
    in step. NOTE: that internal name changes if the field is ever deleted
    and recreated, so if this starts coming back empty, re-check clients'
    user_list field metadata for its current relation id.
    """
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=(
        "SELECT source_table_id, target_table_id "
        "FROM _pq_metadata._pq_rl_1339ee9e")) or []
    links = {}
    for r in rows:
        links.setdefault(to_int(r.get("source_table_id")),
                         set()).add(to_int(r.get("target_table_id")))
    return links


@st.cache_data(ttl=300)
def load_projects():
    """{project_id: {name, client, client_id}} - targets for a new task."""
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT p.id AS project_id, p.name AS project,
               COALESCE(c.id, 0) AS client_id,
               COALESCE(c.name, '-') AS client
        FROM {S}.projects p
        LEFT JOIN {S}.clients c ON c.id = p.client_id
    """) or []
    out = {}
    for r in rows:
        pid = to_int(r.get("project_id"))
        if pid is None:
            continue
        out[pid] = {
            "project": r.get("project") or f"Project {pid}",
            "client": r.get("client") or "-",
            "client_id": to_int(r.get("client_id")) or 0,
        }
    return out


def allowed_task_ids(lookup, links, user_id):
    """The tasks `user_id` may log against: those of clients they are on."""
    return {tid for tid, info in lookup.items()
            if to_int(user_id) in links.get(info["client_id"], set())}


def allowed_project_ids(projects, links, user_id):
    """
    Projects `user_id` may create a task under.

    Same client rule as allowed_task_ids, and for the same reason: a task
    created under a client they are not on would immediately vanish from
    their own task picker, since that is what TASK_CHOICES filters on.
    Creating something you cannot then use is worse than not offering it.
    """
    return {pid for pid, info in projects.items()
            if to_int(user_id) in links.get(info["client_id"], set())}


@st.cache_data(ttl=60)
def load_submissions(user_id):
    dbconn = pq.dbconnect(DW_NAME)
    try:
        rows = dbconn.fetch(DW_NAME, query=f"""
            SELECT id, user_id, week_start_date, status, submitted_at,
                   confirmed_by, confirmed_at
            FROM {S}.timetable_submissions
            WHERE user_id::text = '{int(user_id)}'
        """) or []
    except Exception:
        return {}
    return {str(r.get("id")): r for r in rows}

# =====================================================
# Monthly export view  (SHARED BLOCK - keep byte-identical with
# ts_monthly_view / 12507_ts_monthly_view.py)
#
# Two Data Apps cannot import each other, so this block is duplicated.
# Edit it in one place and copy it to the other.
#
# One view per calendar month: ts_reporting.v_monthly_entries_2026_08 and
# so on. The month window is baked into each view's SQL, so reading one
# needs no date filter - ask for the month by name.
#
# Sourced from the live ts_prod tables on purpose, NOT from
# ts_reporting.fact_timetable: that one is a materialized query table and
# does not reflect writes until its underlying query is re-run. A plain
# view over ts_prod is always current, which is what an export needs.
# =====================================================

VIEW_SCHEMA = "ts_reporting"
VIEW_PREFIX = "v_monthly_entries_"

EXPORT_COLUMNS = [
    ("employee", "Employee"),
    ("entry_day", "Date"),
    ("start_time", "Start"),
    ("duration_hours", "Hours"),
    ("billable", "Billable"),
    ("client", "Client"),
    ("project", "Project"),
    ("task", "Task"),
    ("note", "Note"),
    ("approved", "Validated"),
]


def month_bounds(month_start):
    """(first day, first day of next month) - one view's window."""
    if month_start.month == 12:
        return month_start, date(month_start.year + 1, 1, 1)
    return month_start, date(month_start.year, month_start.month + 1, 1)


def view_name(month_start):
    return f"{VIEW_PREFIX}{month_start.strftime('%Y_%m')}"


def monthly_view_sql(month_start):
    start, end = month_bounds(month_start)
    return f"""
SELECT
    t.id                                      AS entry_id,
    t.user_id::text                           AS user_id,
    COALESCE(u.name, u.email, 'Unknown user') AS employee,
    t.date                                    AS entry_date,
    CAST(t.date AS DATE)                      AS entry_day,
    COALESCE(t.duration, 0)                   AS duration_minutes,
    ROUND(COALESCE(t.duration, 0) / 60.0, 2)  AS duration_hours,
    COALESCE(tk.billable, FALSE)              AS billable,
    COALESCE(c.name, '-')                     AS client,
    COALESCE(p.name, '-')                     AS project,
    COALESCE(tk.name, '-')                    AS task,
    COALESCE(t.internal_description, '')      AS note,
    COALESCE(t.approved, FALSE)               AS approved
FROM {S}.timetable t
LEFT JOIN {S}.users    u  ON u.id::text = t.user_id::text
LEFT JOIN {S}.tasks    tk ON tk.id = t.task_id
LEFT JOIN {S}.projects p  ON p.id = tk.project_id
LEFT JOIN {S}.clients  c  ON c.id = p.client_id
WHERE t.date >= TIMESTAMP '{start.isoformat()} 00:00:00'
  AND t.date <  TIMESTAMP '{end.isoformat()} 00:00:00'
ORDER BY employee, t.date
"""


def ensure_monthly_view(month_start):
    """
    Create this month's view, or update it in place when the SQL changed.
    Idempotent: upsert_query looks the table up by name first and only
    falls back to creating it when it does not exist yet.
    """
    return pq.upsert_query(
        name=view_name(month_start),
        query=monthly_view_sql(month_start),
        schema_name=VIEW_SCHEMA,
        database_name=DW_NAME,
        as_view=True,
    )


def fetch_month(month_start, user_id=None):
    """
    One month's view, ordered employee then date. The view is already
    scoped to the month, so only the employee filter is applied here.
    user_id=None means every employee.
    """
    where = "" if user_id is None else f"WHERE user_id = '{int(user_id)}'"
    dbconn = pq.dbconnect(DW_NAME)
    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT * FROM {VIEW_SCHEMA}.{view_name(month_start)}
        {where}
        ORDER BY employee, entry_date
    """) or []
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    dt = pd.to_datetime(df["entry_date"], errors="coerce")
    df["start_time"] = dt.dt.strftime("%H:%M")
    df["entry_day"] = dt.dt.strftime("%Y-%m-%d")
    return df


def build_workbook(df, sheet_title):
    """The month as .xlsx bytes. openpyxl is available in the runtime."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]                  # Excel's hard limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="053763")
    for col, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(df.to_dict("records") if not df.empty else [], start=2):
        for col, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = row.get(key)
            if key in ("approved", "billable"):
                value = "yes" if value in (True, "true", "True", 1, "1") else "no"
            elif key == "duration_hours":
                value = float(value or 0)
            ws.cell(row=r, column=col, value=value)

    widths = {"employee": 22, "entry_day": 12, "start_time": 8, "duration_hours": 8,
              "billable": 9, "client": 20, "project": 24, "task": 28, "note": 50,
              "approved": 10}
    for col, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 16)

    ws.freeze_panes = "A2"
    if not df.empty:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{len(df) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(month_start, who):
    return f"timesheet_{month_start.strftime('%Y-%m')}_{who}.xlsx"

# =====================================================
# Writes
# =====================================================

def create_task(name, project_id, status, billable, description):
    """
    Create a task and return (task_id, error).

    dbconn.insert does not hand back the new row's id, so the task is read
    back and matched on (project_id, name) - the same re-find ts_mcp_server
    does in get_created_row, with the same failure case when it cannot be
    found again. billable is passed as a real bool on purpose: dbconn puts
    string values through Decimal(), so "false" would 400.
    """
    name = (name or "").strip()
    dbconn = pq.dbconnect(DW_NAME)
    dbconn.insert(DW_NAME, S, "tasks", {
        "name": name,
        "project_id": int(project_id),
        "status": status,
        "billable": bool(billable),
        "description": (description or "").strip(),
    })
    load_task_lookup.clear()
    for tid, info in load_task_lookup().items():
        if (info.get("project_id") == int(project_id)
                and str(info.get("task") or "").strip().lower() == name.lower()):
            return tid, None
    return None, ("The task was created but could not be found again. "
                  "Open the entry dialog once more and pick it from the list.")


def duplicate_task(lookup, project_id, name):
    """ts_prod.tasks is unique on (project_id, name) - mirror that here so
    the picker never shows two identical entries under one project."""
    name = (name or "").strip().lower()
    if not name:
        return False
    return any(info.get("project_id") == int(project_id)
               and str(info.get("task") or "").strip().lower() == name
               for info in lookup.values())


def insert_entry(user_id, task_id, entry_dt, duration_minutes, note):
    dbconn = pq.dbconnect(DW_NAME)
    dbconn.insert(DW_NAME, S, "timetable", {
        "user_id": int(user_id),
        "task_id": int(task_id),
        "date": entry_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "duration": int(duration_minutes),
        "internal_description": (note or "").strip(),
        "external_description": "",
        "date_inserted": datetime.utcnow().strftime("%Y-%m-%d"),
        "approved": None,
        "approved_by": None,
    })
    load_entries.clear()


def update_entry(entry_id, task_id, entry_dt, duration_minutes, note):
    dbconn = pq.dbconnect(DW_NAME)
    dbconn.update(DW_NAME, S, "timetable", int(entry_id), {
        "task_id": int(task_id),
        "date": entry_dt.strftime("%Y-%m-%d %H:%M:%S"),   # keep time component
        "duration": int(duration_minutes),
        "internal_description": (note or "").strip(),
    })
    load_entries.clear()


def delete_entry(entry_id):
    dbconn = pq.dbconnect(DW_NAME)
    dbconn.execute(DW_NAME, query=f"DELETE FROM {S}.timetable WHERE id = {int(entry_id)}")
    load_entries.clear()


def submit_week(user_id, week_start):
    dbconn = pq.dbconnect(DW_NAME)
    key = submission_key(user_id, week_start)
    dbconn.upsert(DW_NAME, S, "timetable_submissions", key, {
        "id": key,
        "user_id": int(user_id),
        "week_start_date": week_start.isoformat(),
        "status": "submitted",
        "submitted_at": datetime.utcnow().isoformat(),
        "confirmed_by": None,
        "confirmed_at": None,
    })
    load_submissions.clear()


def unsubmit_week(user_id, week_start):
    dbconn = pq.dbconnect(DW_NAME)
    key = submission_key(user_id, week_start)
    dbconn.execute(DW_NAME, query=(
        f"DELETE FROM {S}.timetable_submissions "
        f"WHERE id = '{key}' AND status = 'submitted'"))
    load_submissions.clear()


def validate_week(user_id, week_start, validated_by_id, existing_submission):
    """Validate a week. This function IS the contract the other surfaces
    read back, so both halves matter: every timetable entry of the user in
    that week gets approved=true + approved_by (the column names, kept as
    they are), and the submission row becomes status='confirmed'.
    ts_mcp_server refuses writes on a validated entry AND on a validated
    week; this app's grid goes read-only on either. Entries are stamped
    before the submission is, so a failure part-way leaves the week merely
    submitted - still locked, and safe to retry, since re-stamping an
    already-validated entry is a no-op."""
    dbconn = pq.dbconnect(DW_NAME)
    key = submission_key(user_id, week_start)
    week_end = week_start + timedelta(days=6)

    rows = dbconn.fetch(DW_NAME, query=f"""
        SELECT id FROM {S}.timetable
        WHERE user_id::text = '{int(user_id)}'
          AND date >= '{week_start.isoformat()}'
          AND date < '{(week_end + timedelta(days=1)).isoformat()}'
    """) or []
    for row in rows:
        rid = to_int(row.get("id"))
        if rid is not None:
            # A real bool, not "true": ts_prod.timetable.approved is a boolean
            # column and the write proxy tries Decimal() on string values,
            # so "true" comes back as a 400 (decimal.ConversionSyntax).
            dbconn.update(DW_NAME, S, "timetable", rid,
                          {"approved": True, "approved_by": int(validated_by_id)})

    dbconn.upsert(DW_NAME, S, "timetable_submissions", key, {
        "id": key,
        "user_id": int(user_id),
        "week_start_date": week_start.isoformat(),
        "status": "confirmed",
        "submitted_at": (existing_submission or {}).get("submitted_at"),
        "confirmed_by": int(validated_by_id),
        "confirmed_at": datetime.utcnow().isoformat(),
    })
    load_entries.clear()
    load_submissions.clear()

# =====================================================
# Dialogs
# =====================================================

def parse_typed_time(text):
    """
    What someone might type into the start box -> a time, or None.

    Deliberately forgiving about separators and about the leading zero,
    because the fast way to enter 09:15 is "915" and being told that is not
    a time would be absurd. "9" is 09:00; "1342", "13.42", "13h42" and
    "13:42" are all 13:42.
    """
    raw = "".join(str(text).split())
    for sep in (".", ",", ";", "h", "H", "u", "U"):
        raw = raw.replace(sep, ":")
    if raw.isdigit() and len(raw) > 2:          # 915 -> 9:15, 1342 -> 13:42
        raw = f"{raw[:-2]}:{raw[-2:]}"
    parts = raw.split(":")
    try:
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 and parts[1] else 0
    except (ValueError, IndexError):
        return None
    if len(parts) > 2 or not (0 <= hour < 24 and 0 <= minute < 60):
        return None
    return time(hour, minute)


def start_time_input(container, value, key):
    """
    Start time: a quarter-hour menu that also accepts any minute typed.

    st.time_input cannot do both. Its `step` sets which values are VALID,
    not merely which ones the menu lists, so the step short enough to type
    13:42 is also the one that makes the menu 1440 rows long. A selectbox
    that accepts new options separates the two concerns: the menu stays at
    SLOT_MIN, and anything typed is parsed instead of rejected.

    An off-grid value is spliced into the options at its sorted position, so
    an entry starting at 13:42 opens the menu on 13:42 rather than on
    nothing - which is what used to send it back to midnight.

    Where the runtime's Streamlit predates accept_new_options, falls back to
    time_input at START_STEP: still typeable, just a longer menu.
    """
    value = (value or time(9, 0)).replace(second=0, microsecond=0)
    if not CAN_TYPE_NEW_OPTIONS:
        return container.time_input("Start", value=value, step=START_STEP, key=key)

    current = value.strftime("%H:%M")
    choices = sorted({f"{h:02d}:{m:02d}" for h in range(24)
                      for m in range(0, 60, SLOT_MIN)} | {current})
    picked = container.selectbox(
        "Start", choices, index=choices.index(current), key=key,
        accept_new_options=True,
        help="Pick a quarter hour, or type any time - 13:42, 1342 and 13.42 all work",
    )
    typed = parse_typed_time(picked)
    if typed is None:
        # Keep the old value rather than guessing: a start time silently
        # moved is worse than one that refused to move.
        container.caption(f"\"{picked}\" is not a time - keeping {current}.")
        return value
    return typed


def task_selectbox(lookup, key, current=None):
    """Task picker. Nothing preselected unless `current` names a live task -
    a new entry shows the placeholder, so the first task in the list is
    never saved just because nobody touched the box. Returns None while
    the placeholder is showing; callers gate their Save on that."""
    ids = sorted(lookup, key=lambda i: (lookup[i]["client"], lookup[i]["project"], lookup[i]["task"]))
    if TASK_CHOICES is not None:
        ids = [i for i in ids if i in TASK_CHOICES]
    if not ids:
        st.warning("You are not assigned to any client yet, so there is no task "
                   "to log against. Ask an administrator to add you to one.")
        return None
    if current is not None and current not in lookup:
        # Not "not yours" - not known here at all. load_task_lookup is cached
        # for 300s while load_entries is cached for 60s, so a task created on
        # another surface (the MCP) can label its own entry "?" for the four
        # minutes in between. Blaming client access there sends people to an
        # administrator over a cache that is about to expire by itself.
        st.caption("This task was created very recently and this page has not "
                   "picked it up yet - it appears within a few minutes.")
    elif current is not None and current not in ids:
        # Same outcome as ts_mcp_server's update_time_entry, which re-checks
        # client access against the entry's existing task: the entry cannot be
        # saved as it stands, but re-pointing it at one of your own tasks works.
        st.caption("This entry's task belongs to a client you are not assigned "
                   "to. Pick one of your own tasks to save any change.")
    index = ids.index(current) if current in ids else None
    return st.selectbox("Task", ids, index=index, key=key,
                        placeholder="Choose a task...",
                        format_func=lambda i: f"{lookup[i]['client']} - {lookup[i]['task']}")


def existing_task_fields(info):
    """The picked task's own fields, shown but not editable. Editing one
    would change it for everyone logging against it and relabel every
    historical entry, so this is context only - ts_mcp_server exposes no
    task-update tool either."""
    c1, c2 = st.columns([2, 1])
    c1.text_input("Project", value=f"{info['client']} - {info['project']}",
                  disabled=True, key="add_ro_project")
    c2.text_input("Status", value=info["status"], disabled=True, key="add_ro_status")
    st.checkbox("Billable", value=info["billable"], disabled=True, key="add_ro_billable")
    st.text_area("Description", value=info["description"] or "—",
                 disabled=True, height=68, key="add_ro_desc")


def new_task_fields(projects):
    """Blank task form. Returns the entered values, project_id None until
    one is chosen - Save stays disabled while that or the name is empty."""
    ids = sorted(projects, key=lambda i: (projects[i]["client"], projects[i]["project"]))
    if PROJECT_CHOICES is not None:
        ids = [i for i in ids if i in PROJECT_CHOICES]
    if not ids:
        st.info("There is no project you can add a task to yet. Ask an "
                "administrator to assign you to a client.")
        return None

    c1, c2 = st.columns([2, 1])
    project_id = c1.selectbox(
        "Project", ids, index=None, key="add_new_project",
        placeholder="Choose a project...",
        format_func=lambda i: f"{projects[i]['client']} - {projects[i]['project']}")
    status = c2.selectbox("Status", TASK_STATUSES,
                          index=TASK_STATUSES.index(NEW_TASK_STATUS),
                          key="add_new_status")
    name = st.text_input("Task name", key="add_new_name",
                         placeholder="What is this task called?")
    billable = st.checkbox("Billable", value=NEW_TASK_BILLABLE, key="add_new_billable")
    description = st.text_area("Description", key="add_new_desc", height=68,
                               placeholder="Optional")
    return {"project_id": project_id, "status": status, "name": name,
            "billable": billable, "description": description}


def month_options():
    """Every exportable month, newest first: FIRST_EXPORT_MONTH to now."""
    today = date.today()
    m = date(today.year, today.month, 1)
    out = []
    while m >= FIRST_EXPORT_MONTH:
        out.append(m)
        m = date(m.year - 1, 12, 1) if m.month == 1 else date(m.year, m.month - 1, 1)
    return out or [FIRST_EXPORT_MONTH]


def name_slug(user):
    text = str(user_display_name(user)).lower()
    return "".join(ch if ch.isalnum() else "_" for ch in text).strip("_") or "employee"


@st.dialog("Export to Excel")
def export_dialog(can_manage, user_ids, user_by_id, default_month):
    """
    Pick a month and whose entries, build the file, then download it.

    Two steps on purpose: building means asserting the month's view and
    querying it, which is too slow to do on every keystroke, and
    st.download_button needs the bytes in hand before it can be drawn.

    Managers only. An export reaches across every employee and every week,
    including weeks the viewer cannot open in the calendar, so the button
    that opens this is itself behind can_manage - the check below is the
    second lock, not the first.
    """
    if not can_manage:
        st.error("Exporting is limited to managers.")
        return

    months = month_options()
    # The week on screen can sit outside that range - navigate far enough
    # back or forward - so fall back to the newest month rather than
    # offering one the range does not contain.
    default = default_month if default_month in months else months[0]
    export_month = st.selectbox(
        "Month", months, index=months.index(default), key="export_month",
        format_func=lambda d: d.strftime("%B %Y"),
    )
    export_who = st.selectbox(
        "Employees", ["all"] + user_ids, index=0, key="export_who",
        format_func=lambda w: ("All employees" if w == "all"
                               else user_display_name(user_by_id[w])),
    )

    if st.button("Create export", type="primary", use_container_width=True,
                 key="export_build"):
        try:
            with st.spinner("Updating the view and collecting entries..."):
                ensure_monthly_view(export_month)
                target = None if export_who == "all" else int(export_who)
                rows = fetch_month(export_month, target)
                who_label = ("all_employees" if export_who == "all"
                             else name_slug(user_by_id[export_who]))
                st.session_state.export_blob = (
                    (export_month, export_who),        # what these bytes are for
                    export_filename(export_month, who_label),
                    build_workbook(rows, export_month.strftime("%b %Y")),
                    len(rows),
                )
        except Exception as exc:
            st.session_state.export_blob = None
            st.error(f"Export failed: {exc}")

    # Read AFTER the build button so the file is offered in the same run.
    # A file built for other settings is stale the moment they change.
    blob = st.session_state.get("export_blob")
    if not blob or blob[0] != (export_month, export_who):
        return
    _, fname, data, count = blob
    if not count:
        st.info("No entries logged in that month - nothing to export.")
        return
    st.success(f"{count} entries ready.")
    st.download_button(
        f"Download {fname}", data=data, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True, key="export_download",
    )


@st.dialog("New entry")
def add_dialog(user_id, day, lookup, default_start=time(9, 0)):
    st.caption(day.strftime("%A %d %B %Y"))
    task_id = task_selectbox(lookup, "add_task")

    # Picked a task: show what it is. Picked nothing: offer to create one.
    new_task = None
    if task_id is not None:
        existing_task_fields(lookup[task_id])
    else:
        st.caption("No task picked — fill these in to create a new one.")
        new_task = new_task_fields(load_projects())

    st.divider()
    c1, c2 = st.columns(2)
    start = start_time_input(c1, default_start, "add_start_pick")
    dur = c2.number_input("Duration (min)", min_value=5, step=15, value=60, key="add_dur")
    note = st.text_area("Note", key="add_note", height=80)

    ready = task_id is not None or bool(
        new_task and new_task["project_id"] is not None and new_task["name"].strip())
    label = "Save" if task_id is not None else "Create task and save"
    if st.button(label, type="primary", use_container_width=True, key="add_save",
                 disabled=not ready):
        error = None
        if task_id is None:
            if duplicate_task(lookup, new_task["project_id"], new_task["name"]):
                error = ("That project already has a task with this name. "
                         "Pick it from the list above instead.")
            else:
                task_id, error = create_task(
                    new_task["name"], new_task["project_id"], new_task["status"],
                    new_task["billable"], new_task["description"])
        if error:
            st.error(error)
        else:
            insert_entry(user_id, task_id, datetime.combine(day, start), dur, note)
            st.rerun()


@st.dialog("Entry")
def entry_dialog(entry, lookup, editable):
    info = lookup.get(to_int(entry.get("task_id")), LOOKUP_DEFAULT)
    dt = entry["dt"]
    st.markdown(
        f"**{info['task']}**  \n"
        f"{info['client']} - {info['project']}  \n"
        f"{dt.strftime('%a %d %b')} - {dt.strftime('%H:%M')} to "
        f"{(dt + timedelta(minutes=int(entry['duration']))).strftime('%H:%M')} "
        f"({fmt_dur(entry['duration'])})"
    )
    if entry.get("internal_description"):
        st.caption(entry["internal_description"])

    if not editable:
        st.info("This entry is read-only (week submitted/validated, entry validated, or not your calendar).")
        return

    with st.form("edit_form", border=False):
        task_id = task_selectbox(lookup, "edit_task", current=to_int(entry.get("task_id")))
        c1, c2, c3 = st.columns(3)
        new_day = c1.date_input("Day", value=dt.date(), key="edit_day")
        new_start = start_time_input(c2, dt.time(), "edit_start_pick")
        new_dur = c3.number_input("Duration (min)", min_value=5, step=15,
                                  value=int(entry["duration"]), key="edit_dur")
        new_note = st.text_area("Note", value=entry.get("internal_description") or "",
                                key="edit_note", height=80)
        s1, s2 = st.columns(2)
        if s1.form_submit_button("Save", type="primary", use_container_width=True,
                                 disabled=task_id is None):
            update_entry(entry["id"], task_id, datetime.combine(new_day, new_start), new_dur, new_note)
            st.rerun()
        if s2.form_submit_button("Delete", use_container_width=True):
            delete_entry(entry["id"])
            st.rerun()


@st.dialog("Submit week")
def submit_dialog(user_id, week_start, day_totals):
    week_end = week_start + timedelta(days=6)
    st.write(f"Submit **{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}** for validation?")
    shortfalls = []
    for i in WORKDAYS:
        total = day_totals.get(week_start + timedelta(days=i), 0)
        if total >= DAY_TARGET_MIN:
            st.markdown(f"{DAY_ABBREV[i]} - {fmt_dur(total)} - ok")
        else:
            shortfalls.append(i)
            st.markdown(f"{DAY_ABBREV[i]} - {fmt_dur(total)} - **missing {fmt_dur(DAY_TARGET_MIN - total)}**")
    st.caption("Submitting locks the week for editing until it is unsubmitted or validated.")
    if not shortfalls:
        if st.button("Submit week", type="primary", use_container_width=True, key="submit_ok"):
            submit_week(user_id, week_start)
            st.rerun()
    else:
        st.warning("Not every workday reaches the 8h minimum yet.")
        if st.button("Submit anyway", use_container_width=True, key="submit_anyway"):
            submit_week(user_id, week_start)
            st.rerun()


@st.dialog("Validate week")
def validate_dialog(target_user, week_start, validated_by_id, existing_submission):
    week_end = week_start + timedelta(days=6)
    st.write(
        f"Validate the week of **{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}** "
        f"for **{user_display_name(target_user)}**?"
    )
    st.caption("This validates every entry in the week and marks the submission "
               "as confirmed. Validated entries can no longer be edited by anyone.")
    if st.button("Validate week", type="primary", use_container_width=True, key="validate_ok"):
        validate_week(to_int(target_user.get("id")), week_start, validated_by_id, existing_submission)
        st.rerun()

# =====================================================
# Grid model builders (pure functions -> unit-testable)
# =====================================================

def assign_lanes(evs):
    """Greedy lanes for overlapping events -> [(event, lane, n_lanes)]."""
    out = []                        # [event, lane, start_min, end_min]
    for e in sorted(evs, key=lambda e: (e["dt"], -int(e["duration"]))):
        s = e["dt"].hour * 60 + e["dt"].minute
        en = s + max(int(e["duration"]), 15)
        taken = {o[1] for o in out if o[2] < en and o[3] > s}
        out.append([e, next(i for i in range(len(taken) + 1) if i not in taken), s, en])
    # n_lanes is only known once every overlapping event has a lane
    return [(e, lane, max(o[1] for o in out if o[2] < en and o[3] > s) + 1)
            for e, lane, s, en in out]


def build_grid_frames(week_df, days, lookup):
    """Split week entries into the chart dataframe (timed) and the
    no-start-time list. Returns (bars_df, untimed_list, start_h, end_h)."""
    timed, untimed = {}, []
    if not week_df.empty:
        for _, r in week_df.iterrows():
            e = r.to_dict()
            if e["dt"].time() == time(0, 0):
                untimed.append(e)
            else:
                timed.setdefault(e["dt"].date(), []).append(e)

    start_h, end_h = GRID_START_H, GRID_END_H
    for evs in timed.values():
        for e in evs:
            endt = e["dt"] + timedelta(minutes=int(e["duration"]))
            eh = endt.hour + (1 if (endt.minute or endt.second) else 0)
            if endt.date() != e["dt"].date():
                eh = 24
            start_h = min(start_h, e["dt"].hour)
            end_h = max(end_h, min(eh, 24))
    start_h = max(0, start_h)
    end_h = min(24, max(end_h, start_h + 1))

    rows = []
    for i, d in enumerate(days):
        for e, lane, n_lanes in assign_lanes(timed.get(d, [])):
            info = lookup.get(to_int(e.get("task_id")), LOOKUP_DEFAULT)
            start = e["dt"].hour + e["dt"].minute / 60
            dur_h = max(int(e["duration"]), 15) / 60
            locked = is_true(e.get("approved"))
            end_t = (e["dt"] + timedelta(minutes=int(e["duration"]))).strftime("%H:%M")
            rows.append({
                "entry_id": int(e["id"]),
                "day_i": i,
                "lane": lane,
                "n_lanes": n_lanes,
                "y0": start,
                "y1": min(start + dur_h, 24),
                "color": info["color"],
                "label": f"{e['dt'].strftime('%H:%M')} {info['client']}",
                "client": info["client"],
                "task": info["task"],
                "tijd": f"{e['dt'].strftime('%H:%M')} - {end_t}",
                "duur": fmt_dur(e["duration"]),
                "locked": locked,
                "note": (e.get("internal_description") or "")[:120],
            })
    return pd.DataFrame(rows), untimed, start_h, end_h


def occupied_slots(bars_df):
    """
    {day_index: {slot, ...}} for every slot an entry overlaps, where a slot
    is SLOT_MIN minutes counted from midnight: slot 37 is 09:15.

    Rounded outwards, so an entry from 09:20 to 09:50 occupies 09:15 and
    09:30 - a partly covered slot is not offered as free.
    """
    busy = {}
    if bars_df is None or bars_df.empty:
        return busy
    for r in bars_df.itertuples():
        lo = round(float(r.y0) * 60) // SLOT_MIN            # floor
        hi = -(-round(float(r.y1) * 60) // SLOT_MIN)        # ceil
        busy.setdefault(int(r.day_i), set()).update(range(lo, max(hi, lo + 1)))
    return busy


def build_cells_frame(days, start_h, end_h, bars_df=None):
    """Click targets for ADDING an entry: only the SLOT_MIN-minute blocks
    that are still free. Slots covered by an existing entry are left out,
    so a click on a block can never fall through to the add dialog and
    create an overlap."""
    busy = occupied_slots(bars_df)
    rows = []
    for i, d in enumerate(days):
        taken = busy.get(i, set())
        for slot in range(start_h * SLOTS_PER_H, end_h * SLOTS_PER_H):
            if slot in taken:
                continue
            h, m = divmod(slot * SLOT_MIN, 60)
            rows.append({"day_i": i,
                         "cell_day": d.isoformat(), "cell_hour": h, "cell_min": m,
                         "y0": slot / SLOTS_PER_H,
                         "hint": f"new entry at {h:02d}:{m:02d}"})
    return pd.DataFrame(rows, columns=["day_i", "cell_day", "cell_hour",
                                       "cell_min", "y0", "hint"])


def selected_customdata(event):
    """Pull customdata lists from a st.plotly_chart on_select event,
    preferring entry clicks over cell clicks when both are hit."""
    points = (event.get("selection") or {}).get("points") or []
    datas = [list(p["customdata"]) for p in points if p.get("customdata")]
    # an entry bar wins over the cell under it
    return next((cd for cd in datas if cd[0] == "entry"), datas[0] if datas else None)

# =====================================================
# Google login (OAuth 2.0 authorization code flow)
# =====================================================
#
# Access is granted only to visitors who complete a Google login whose
# verified email address matches a row in ts_prod.users. Nothing renders
# and no timesheet query runs before that check passes, so publishing the
# app as "Public" exposes the login page and nothing else.
#
# Single-app flow, no callback handler: the login page links to Google's
# consent screen with redirect_uri = this app's own published URL. Google
# sends the browser back there with ?code=&state=, this script exchanges
# the code at Google's token endpoint using the client secret from the
# Secret Store, reads the identity out of the returned id_token, and
# keeps it in an encrypted cookie for SESSION_HOURS so a refresh does not
# restart the login.
#
# The redirect leaves and re-enters the app, so st.session_state is gone
# by the time the code comes back. The CSRF state and the OIDC nonce are
# therefore self-verifying - HMAC-signed with the cookie password and
# timestamped - rather than stored server-side. The state is deliberately
# NOT also pinned in a cookie: writing a cookie on every login-page
# render would round-trip the cookie component on each rerun, and a
# stolen code is already useless without our client secret.

GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_ISSUERS = ("accounts.google.com", "https://accounts.google.com")

# Must match an "Authorized redirect URI" on the Google Cloud OAuth
# client byte for byte, trailing slash included.
PUBLISHED_APP_URL = (
    "https://app.eu.peliqan.io/apps/"
    "dkV4ZE1JMW5obnhsblFJemM5anhKZEQ5UTZYWVp6TTNLZmhPRDJEcXZxeDljcnBndTBWcndnaWpIVmRoYjJwaw==/"
)
# For local `streamlit run` development, point this at http://localhost:8501/
# and register that as a second redirect URI on the same Google client.
REDIRECT_URI = PUBLISHED_APP_URL

SECRET_CLIENT_ID = "google_login_client_id"
SECRET_CLIENT_SECRET = "google_login_client_secret"
SECRET_COOKIE_PASSWORD = "ts_cookie_password"

SESSION_HOURS = 12            # how long one login stays valid
STATE_MAX_AGE = 30 * 60       # a started login must be completed within this
COOKIE_PREFIX = "ts_my_week_"
COOKIE_SESSION = "session"


class LoginError(Exception):
    """Anything that must send the visitor back to the login page."""


class StaleCodeError(LoginError):
    """
    The authorization code was already redeemed, or it expired.

    Not a misconfiguration and not worth an error banner: the visitor
    simply gets the sign-in page back. Google's redirect lands on the
    top-level window (the login uses target="_top"), so ?code= lives on
    in a URL this app cannot rewrite - every refresh replays it.
    """


def now_ts():
    # NOTE: `time` is datetime.time in this module, so no `import time`.
    return int(datetime.now(timezone.utc).timestamp())


@st.cache_data(ttl=3600)
def oauth_config():
    return {
        "client_id": pq.get_secret(SECRET_CLIENT_ID),
        "client_secret": pq.get_secret(SECRET_CLIENT_SECRET),
        "cookie_password": pq.get_secret(SECRET_COOKIE_PASSWORD),
    }


def b64url_encode(raw):
    return base64.urlsafe_b64encode(raw).decode().rstrip("=")


def b64url_decode(txt):
    return base64.urlsafe_b64decode(str(txt) + "=" * (-len(str(txt)) % 4))


def sign(payload):
    key = oauth_config()["cookie_password"].encode()
    return b64url_encode(hmac.new(key, payload.encode(), hashlib.sha256).digest())


def make_state():
    """Signed, timestamped state; its random part also seeds the nonce."""
    rand = secrets.token_urlsafe(16)
    body = b64url_encode(json.dumps({"r": rand, "t": now_ts()}).encode())
    return body + "." + sign(body), rand


def read_state(state):
    """Verify our own signature + age, return the state's random part."""
    try:
        body, mac = str(state).split(".", 1)
    except (ValueError, AttributeError):
        raise LoginError("The login response carried no state - please sign in again.")
    if not hmac.compare_digest(mac, sign(body)):
        raise LoginError("The login response did not belong to a login started here.")
    try:
        data = json.loads(b64url_decode(body))
    except Exception:
        raise LoginError("The login state was unreadable - please sign in again.")
    if now_ts() - int(data.get("t") or 0) > STATE_MAX_AGE:
        raise LoginError("This login took too long to complete - please sign in again.")
    return str(data.get("r") or "")


def nonce_for(state_rand):
    return sign("nonce|" + state_rand)


def build_auth_url(state, state_rand):
    return GOOGLE_AUTH_URL + "?" + urllib.parse.urlencode({
        "client_id": oauth_config()["client_id"],
        "redirect_uri": REDIRECT_URI,
        "response_type": "code",
        "scope": "openid email profile",
        "state": state,
        "nonce": nonce_for(state_rand),
        "access_type": "online",
        "prompt": "select_account",
    })


def exchange_code(code):
    cfg = oauth_config()
    body = urllib.parse.urlencode({
        "code": code,
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "redirect_uri": REDIRECT_URI,
        "grant_type": "authorization_code",
    }).encode()
    req = urllib.request.Request(
        GOOGLE_TOKEN_URL, data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            detail = json.loads(exc.read().decode("utf-8")).get("error", "")
        except Exception:
            pass
        if detail == "invalid_grant":
            raise StaleCodeError("This authorization code was already used or has expired.")
        raise LoginError(
            f"Google refused this login ({detail or exc.code}). Usually the "
            "redirect URI or the client secret does not match the Google Cloud client."
        )
    except Exception:
        raise LoginError("Could not reach Google to finish the login - please try again.")


def read_identity(tokens, expected_nonce):
    """
    Validate the id_token's claims. Its signature is not re-checked: the
    token arrived straight from Google's token endpoint over TLS, in
    response to a request carrying our client secret - the one case
    Google documents as not requiring local signature verification.
    """
    id_token = tokens.get("id_token")
    if not id_token:
        raise LoginError("Google returned no identity token.")
    try:
        claims = json.loads(b64url_decode(id_token.split(".")[1]))
    except Exception:
        raise LoginError("The identity token from Google was unreadable.")

    if claims.get("iss") not in GOOGLE_ISSUERS:
        raise LoginError("That identity token was not issued by Google.")
    if claims.get("aud") != oauth_config()["client_id"]:
        raise LoginError("That identity token was issued for a different application.")
    if int(claims.get("exp") or 0) <= now_ts():
        raise LoginError("That identity token had already expired.")
    if expected_nonce and claims.get("nonce") != expected_nonce:
        raise LoginError("That identity token does not match this login attempt.")
    if not is_true(claims.get("email_verified")):
        raise LoginError("This Google account has no verified email address.")

    email = str(claims.get("email") or "").strip().lower()
    if not email:
        raise LoginError("Google returned no email address for this account.")
    return {"email": email, "name": claims.get("name") or email}


def init_cookies():
    """
    Encrypted cookie jar, so a page refresh does not restart the login.
    Optional by design: if the component is missing the app still works,
    the login then lasts only as long as the Streamlit session.
    """
    if st.session_state.get("cookies_unavailable"):
        return None
    try:
        st.cache = st.cache_data     # streamlit_cookies_manager still expects st.cache
        from streamlit_cookies_manager import EncryptedCookieManager
        return EncryptedCookieManager(prefix=COOKIE_PREFIX,
                                      password=oauth_config()["cookie_password"])
    except Exception:
        st.session_state.cookies_unavailable = True
        return None


def store_session(cookies, email, name):
    auth = {"email": email, "name": name, "exp": now_ts() + SESSION_HOURS * 3600}
    st.session_state.auth = auth
    if cookies is not None:
        body = b64url_encode(json.dumps(auth).encode())
        try:
            cookies[COOKIE_SESSION] = body + "." + sign(body)
            cookies.save()
        except Exception:
            pass
    return auth


def read_session(cookies):
    """Live session, else the cookie; None when absent or expired."""
    auth = st.session_state.get("auth")
    if not auth and cookies is not None:
        raw = cookies.get(COOKIE_SESSION)
        if raw:
            try:
                body, mac = str(raw).split(".", 1)
                if hmac.compare_digest(mac, sign(body)):
                    auth = json.loads(b64url_decode(body))
            except Exception:
                auth = None
    if not auth or int(auth.get("exp") or 0) <= now_ts():
        return None
    st.session_state.auth = auth
    return auth


def end_session(cookies):
    st.session_state.pop("auth", None)
    if cookies is not None:
        try:
            cookies[COOKIE_SESSION] = ""
            cookies.save()
        except Exception:
            pass


def login_page(message=None):
    """
    Render the sign-in page and stop the script - never returns.

    The message may also arrive through session_state: mutating query
    params can rerun the script before this renders, so failures stash
    their reason there and it is picked up on the next run.
    """
    stashed = st.session_state.pop("login_message", None)   # pop even if unused
    message = message or stashed
    state, state_rand = make_state()
    st.title("My week")
    if message:
        st.error(message)
    st.write("Sign in with the Google account of your timesheet user to open your week.")
    # target="_top" so the whole browser window follows the redirect even
    # when the app is embedded - Google refuses to render in an iframe.
    st.markdown(
        f"<a href='{build_auth_url(state, state_rand)}' target='_top' "
        "style='display:inline-block;padding:0.55rem 1.1rem;border-radius:0.5rem;"
        "background:#053763;color:#fff;text-decoration:none;font-weight:600;'>"
        "Sign in with Google</a>",
        unsafe_allow_html=True,
    )
    if st.session_state.get("cookies_unavailable"):
        st.caption("Cookies are unavailable here, so a page refresh will ask you to sign in again.")
    st.stop()


# Keys Google appends to the redirect. Only these are removed from the
# URL after a login - a blanket st.query_params.clear() would also drop
# Peliqan's own parameters (embed=true, session_token).
GOOGLE_PARAMS = ("code", "state", "scope", "authuser", "prompt", "hd", "error", "error_subtype")


def clear_login_params():
    for key in GOOGLE_PARAMS:
        if key in st.query_params:
            del st.query_params[key]


def match_login_to_user(email):
    for u in load_users():
        if str(u.get("email") or "").strip().lower() == email:
            return u
    return None

# ---- the gate: nothing below here runs unauthenticated ----

try:
    oauth_config()
except Exception:
    st.error(
        "Google login is not configured yet. Add these three entries to the "
        f"Peliqan Secret Store: {SECRET_CLIENT_ID}, {SECRET_CLIENT_SECRET} "
        f"and {SECRET_COOKIE_PASSWORD}."
    )
    st.stop()

cookies = init_cookies()
if cookies is not None and not cookies.ready():
    st.stop()                    # cookie component still initialising

auth = read_session(cookies)

if auth is None:
    params = st.query_params
    if params.get("error"):
        denied = f"Google did not complete the login ({str(params.get('error'))})."
        st.session_state.login_message = denied
        clear_login_params()
        login_page(denied)
    code = params.get("code")
    if not code:
        login_page()
    # Spend a code once per session: a rerun must not resubmit it.
    # Reset by a hard refresh, which starts a fresh session - the
    # StaleCodeError branch below is what covers that case.
    consumed = st.session_state.setdefault("consumed_codes", set())
    if str(code) in consumed:
        clear_login_params()
        login_page()
    consumed.add(str(code))
    try:
        state_rand = read_state(params.get("state"))
        identity = read_identity(exchange_code(code), nonce_for(state_rand))
    except StaleCodeError:
        clear_login_params()     # leftover ?code=, not a failed login
        login_page()
    except LoginError as exc:
        st.session_state.login_message = str(exc)
        clear_login_params()
        login_page(str(exc))
    # Session first, then drop ?code= from our own URL. That is all
    # this app can reach: the copy on the top-level window survives,
    # so a replay is caught by StaleCodeError rather than prevented.
    auth = store_session(cookies, identity["email"], identity["name"])
    clear_login_params()
    st.rerun()

elif "code" in st.query_params:
    clear_login_params()    # already signed in: drop a stale code from the URL

login_user = match_login_to_user(auth["email"])
if login_user is None:
    st.title("My week")
    st.error(
        f"{auth['email']} is not a timesheet user. Ask an administrator to add this "
        "email address to the users table, or sign in with another account."
    )
    if st.button("Sign in with another account", key="switch_account"):
        end_session(cookies)
        st.rerun()
    st.stop()

LOGIN_USER_ID = to_int(login_user.get("id"))

# =====================================================
# State
# =====================================================

if "week_start" not in st.session_state:
    st.session_state.week_start = monday_of(date.today())
if "cal_nonce" not in st.session_state:
    st.session_state.cal_nonce = 0
if "pending" not in st.session_state:
    st.session_state.pending = None

# =====================================================
# Who am I / whose calendar am I viewing
# =====================================================

users = load_users()
if not users:
    st.warning("No users found - check ts_prod.users.")
    st.stop()

user_by_id = {to_int(u.get("id")): u for u in users}
user_ids = list(user_by_id.keys())

# One row for everything at the top, bottom-aligned so the buttons sit on
# the same baseline as the inputs beside them rather than level with their
# labels. Weights sum to 10.0; gap_col is deliberately left empty,
# separating the filters from the actions on the right.
#
# The last TWO weights are 2.0 each and the header row below ends with the
# same pair, which is what stacks the four action buttons into an aligned
# 2x2 block: Export over Validate, the account over Submit. Change one row
# and you must change the other.
view_col, week_col, gap_col, export_col, acct_col = st.columns(
    [2.2, 1.6, 2.2, 2.0, 2.0], vertical_alignment="bottom")

# The viewer is whoever logged in with Google - there is no other identity.
viewer_id = LOGIN_USER_ID
viewer = user_by_id[viewer_id]
viewer_scope = to_int(viewer.get("scope")) or 1
can_manage = viewer_scope >= MANAGE_SCOPE

if can_manage:
    viewing_id = view_col.selectbox(
        "Viewing calendar of", user_ids,
        index=user_ids.index(viewer_id),
        key=f"viewing_{viewer_id}",     # keyed per viewer: switching viewer resets to self
        format_func=lambda i: user_display_name(user_by_id[i]),
    )
else:
    viewing_id = viewer_id              # scope 1: own calendar only

viewing_user = user_by_id[viewing_id]
is_self = viewing_id == viewer_id

# The account collapses into a popover: a strip of its own above the
# filters cost a whole row of height for two small things. The name is the
# label, the address and Sign out live inside.
with acct_col.popover(auth["name"], use_container_width=True):
    st.caption(auth["email"])
    if st.button("Sign out", key="sign_out", use_container_width=True):
        end_session(cookies)
        st.rerun()

# Exporting spans months, so it sits with the filters rather than anywhere
# in the week grid. Managers only - scope 1 never sees the button.
# Everything else happens in the dialog.
if can_manage and export_col.button("Export to Excel", key="open_export",
                                    use_container_width=True):
    export_dialog(can_manage, user_ids, user_by_id,
                  date(st.session_state.week_start.year,
                       st.session_state.week_start.month, 1))


def go_to_week(new_start):
    st.session_state.week_start = new_start
    st.rerun()


# Week navigation by date: pick any day, land on that week (Monday-based).
# The picker's key contains the current week: Streamlit forbids writing a
# widget's state after instantiation, so instead of updating the widget
# after button navigation we let it be rebuilt with the new default.
picked_day = week_col.date_input(
    "Show week of", value=st.session_state.week_start,
    key=f"week_picker_{st.session_state.week_start.isoformat()}",
    help="Pick any date to jump to that week",
)
if isinstance(picked_day, (list, tuple)):
    picked_day = picked_day[0] if picked_day else st.session_state.week_start
if picked_day and monday_of(picked_day) != st.session_state.week_start:
    go_to_week(monday_of(picked_day))

# =====================================================
# Data for the viewed user + permissions for this week
# =====================================================

entries = load_entries(viewing_id)
lookup = load_task_lookup()

# A write from another surface - the MCP - is invisible here until this
# app's caches expire, and they do not expire together: entries are cached
# for 60s, tasks/projects/client links for 300s. Every .clear() elsewhere in
# this file follows one of THIS app's own writes, so nothing here ever hears
# about anyone else's. An entry whose task_id is missing from the lookup is
# the only signal available that someone else has written, so treat it as
# one: refresh the three slow caches and rerun. Otherwise the entry renders
# as "?" and its task cannot be picked for the four minutes in between.
#
# Links and projects go too, not just tasks: a task new enough to be missing
# usually arrives with a new project and client, and refreshing the lookup
# alone would only move the entry from "?" to "not one of your clients".
#
# Once per id per session. An entry can also point at a task that is
# genuinely gone, and that one never resolves - remembering which ids have
# already been retried is what stops this rerunning forever.
if not entries.empty:
    retried = st.session_state.setdefault("refreshed_task_ids", set())
    unknown = {t for t in (to_int(v) for v in entries["task_id"])
               if t is not None and t not in lookup} - retried
    if unknown:
        retried |= unknown
        load_task_lookup.clear()
        load_client_links.clear()
        load_projects.clear()
        st.rerun()

# Employees may only log against clients they are assigned to, exactly as
# ts_mcp_server's _check_entry_client_access limits log_time_entry; managers
# and admins may log against anything. Keyed on the VIEWED user, since that
# is whose entry would be written - though only they can write it.
TASK_CHOICES = None if can_manage else allowed_task_ids(
    lookup, load_client_links(), viewing_id)
# Creating a task is open to every scope, but the project list follows the
# same client rule as the task list - see allowed_project_ids.
PROJECT_CHOICES = None if can_manage else allowed_project_ids(
    load_projects(), load_client_links(), viewing_id)
submissions = load_submissions(viewing_id)

week_start = st.session_state.week_start
week_end = week_start + timedelta(days=6)
days = [week_start + timedelta(days=i) for i in range(7)]

submission = submissions.get(submission_key(viewing_id, week_start))
wk_status = (submission or {}).get("status")
is_locked_week = wk_status in ("submitted", "confirmed")
is_confirmed = wk_status == "confirmed"

can_edit = is_self and not is_locked_week   # add/edit/delete entries, and submit
can_unsubmit = is_self and wk_status == "submitted"
can_validate = can_manage and wk_status == "submitted"

week_df = entries[(entries["dt"].dt.date >= week_start) & (entries["dt"].dt.date <= week_end)] if not entries.empty else entries
day_totals = {}
if not week_df.empty:
    day_totals = week_df.groupby(week_df["dt"].dt.date)["duration"].sum().to_dict()

# ---- open a dialog queued by a previous click (then clear the flag;
#      interactions inside the dialog rerun only the dialog fragment) ----
pending, st.session_state.pending = st.session_state.pending, None
if pending:
    kind = pending[0]
    if kind == "entry" and not entries.empty:
        match = entries[entries["id"].astype(str) == str(pending[1])]
        if not match.empty:
            e = match.iloc[0].to_dict()
            entry_dialog(e, lookup, editable=can_edit and not is_true(e.get("approved")))
    elif kind == "add" and can_edit:
        # len check: a pending queued before slots existed carries no minute
        minute = pending[3] if len(pending) > 3 else 0
        add_dialog(viewing_id, pending[1], lookup,
                   default_start=time(pending[2], minute))

# =====================================================
# Header
# =====================================================

# Weights sum to 10.0, matching the filter row above. title_col absorbs
# whatever the buttons beside it do not use.
# Ends with the same 2.0 + 2.0 pair as the filter row above, so Validate
# lands under Export and Submit under the account. title_col gives up the
# width that second column costs.
prev_col, next_col, this_col, title_col, validate_col, submit_col = st.columns(
    [1.15, 1.0, 1.0, 2.85, 2.0, 2.0])

if prev_col.button("Previous week", key="prev_week", use_container_width=True):
    go_to_week(week_start - timedelta(days=7))
if next_col.button("Next week", key="next_week", use_container_width=True):
    go_to_week(week_start + timedelta(days=7))
this_week = monday_of(date.today())
if week_start != this_week:
    if this_col.button("This week", key="this_week", use_container_width=True):
        go_to_week(this_week)

iso_week = week_start.isocalendar()[1]
whose = "my week" if is_self else user_display_name(viewing_user)
title_col.markdown(
    f"### {week_start.strftime('%d %b')} - {week_end.strftime('%d %b %Y')} "
    f"<span style='color:#70757a;font-size:0.8rem;'>week {iso_week} - {whose}</span>",
    unsafe_allow_html=True,
)

# Left column is always the validate action, right column always the
# submit/unsubmit one, so a button never moves sideways as the week
# changes state - it only appears or disappears.
if is_confirmed:
    by = user_by_id.get(to_int((submission or {}).get("confirmed_by")))
    validate_col.success("Week validated" + (f" by {user_display_name(by)}" if by else ""))
elif wk_status == "submitted":
    if can_validate and validate_col.button("Validate", key="validate_btn",
                                            type="primary", use_container_width=True):
        validate_dialog(viewing_user, week_start, viewer_id, submission)
    if can_unsubmit and submit_col.button("Unsubmit", key="unsubmit_btn",
                                          use_container_width=True):
        unsubmit_week(viewing_id, week_start)
        st.rerun()
elif can_edit:
    if submit_col.button("Submit week", key="submit_week_btn",
                         type="primary", use_container_width=True):
        submit_dialog(viewing_id, week_start, day_totals)
else:
    submit_col.caption("Week not submitted yet.")

# per-day 8h minimum summary (replaces the old weekly progress bar:
# the target is a minimum of 8h per workday, not a 40h weekly pool)
days_ok = sum(1 for i in WORKDAYS if day_totals.get(week_start + timedelta(days=i), 0) >= DAY_TARGET_MIN)
week_total = int(sum(day_totals.values()))
st.caption(f"{days_ok} of {len(WORKDAYS)} workdays at the 8h minimum - {fmt_dur(week_total)} logged in total")

# day totals row, aligned with the 7 chart columns. Each day carries its
# own "+" button (replaces the old global New button), so adding an entry
# always starts from an explicit day.
tot_cols = st.columns(7)
for i, d in enumerate(days):
    total = int(day_totals.get(d, 0))
    if total == 0:
        color, txt = "#b0b3b8", "0h"
    elif i in WORKDAYS and total < DAY_TARGET_MIN:
        color, txt = "#ea8600", fmt_dur(total)
    else:
        color, txt = "#188038", fmt_dur(total)
    name_color = "#1a73e8" if d == date.today() else "#3c4043"
    with tot_cols[i]:
        st.markdown(
            f"<div style='text-align:center;font-size:0.78rem;'>"
            f"<b style='color:{name_color};'>{DAY_ABBREV[i]} {d.day}</b>"
            f"<span style='color:{color};'> - {txt}</span></div>",    
            unsafe_allow_html=True,
        )
        if can_edit:
            if st.button("Add an entry", key=f"add_day_{d.isoformat()}",
                         help=f"Add an entry on {d.strftime('%a %d %b')}",
                         use_container_width=True):
                add_dialog(viewing_id, d, lookup)
        

# =====================================================
# The calendar chart (Plotly with Altair/Vega styling)
# =====================================================

bars_df, untimed, start_h, end_h = build_grid_frames(week_df, days, lookup)
# free hours only: an add target under an entry would let a click on a
# block fall through and create an overlapping entry
cells_df = build_cells_frame(days, start_h, end_h, bars_df)

fig = go.Figure()

# Invisible click targets for empty hour cells (behind the entry bars).
#
# Bars, not scatter markers. A marker is sized in pixels, so a 34px square
# covered only the middle ~18% of a day column and ~65% of a row - while
# Plotly's "closest" hover still turned the cursor into a pointer well
# outside it. That is how you got a pointing finger over places with no
# add target: no tooltip, and a click that selected nothing. A bar lives
# in data coordinates, so width=1 and height=1 tile the cell exactly, at
# any window size, and hover only fires inside it.
if can_edit and not cells_df.empty:
    fig.add_trace(go.Bar(
        x=cells_df["day_i"],
        width=1.0,                         # exactly one day column
        base=cells_df["y0"],
        y=[1 / SLOTS_PER_H] * len(cells_df),   # exactly one SLOT_MIN block
        marker=dict(color="rgba(76,120,168,0.001)", line=dict(width=0)),
        customdata=[["add", r.cell_day, int(r.cell_hour), int(r.cell_min)]
                    for r in cells_df.itertuples()],
        hovertext=cells_df["hint"],
        hovertemplate="%{hovertext}<extra></extra>",
        textposition="none",               # the hint is a tooltip, not a label
        showlegend=False,
        name="cells",
    ))

# entry bars: base = start hour, length = duration; lanes share the column
if not bars_df.empty:
    lane_w = 0.9 / bars_df["n_lanes"].astype(float)
    x_pos = bars_df["day_i"] - 0.45 + (bars_df["lane"] + 0.5) * lane_w
    fig.add_trace(go.Bar(
        x=x_pos,
        width=(lane_w - 0.03).clip(lower=0.08),
        base=bars_df["y0"],
        y=bars_df["y1"] - bars_df["y0"],
        marker=dict(color=bars_df["color"], opacity=0.92,
                    line=dict(color="white", width=1)),
        text=bars_df["label"],
        textposition="inside",
        insidetextanchor="start",
        textangle=0,
        textfont=dict(color="white", size=10,
                      family="Source Sans Pro, Helvetica Neue, sans-serif"),
        customdata=[["entry", int(i)] for i in bars_df["entry_id"]],
        showlegend=False,
        name="entries",
    ))
    fig.data[-1].hovertemplate = "%{hovertext}<extra></extra>"
    fig.data[-1].hovertext = [
        f"<b>{r.client}</b> - {r.task}<br>{r.tijd} ({r.duur})"
        + ("<br>validated (locked)" if r.locked else "")
        + (f"<br><i>{r.note}</i>" if r.note else "")
        for r in bars_df.itertuples()
    ]

# today's column shading
for i, d in enumerate(days):
    if d == date.today():
        fig.add_vrect(x0=i - 0.5, x1=i + 0.5, fillcolor="#eaf1fb",
                      opacity=0.55, layer="below", line_width=0)

fig.update_layout(
    height=(end_h - start_h) * ROW_PX,
    margin=dict(l=8, r=8, t=30, b=8),
    plot_bgcolor="white",
    paper_bgcolor="rgba(0,0,0,0)",
    barmode="overlay",
    barcornerradius=4,                     # the rounded Altair look
    dragmode=False,
    clickmode="event+select",
    font=dict(family="Source Sans Pro, Helvetica Neue, sans-serif"),
    xaxis=dict(
        range=[-0.5, 6.5],
        tickvals=list(range(7)),
        ticktext=[f"{DAY_ABBREV[i]} {d.day}" for i, d in enumerate(days)],
        side="top", fixedrange=True, showgrid=False, zeroline=False,
        showline=False, ticks="",
        tickfont=dict(size=11, color="#6e6e78"),
    ),
    yaxis=dict(
        range=[end_h, start_h],            # morning on top
        tickvals=list(range(start_h, end_h + 1)),
        ticktext=[f"{h:02d}:00" for h in range(start_h, end_h + 1)],
        fixedrange=True, zeroline=False, showline=False, ticks="",
        gridcolor="#eeeeef", gridwidth=1,
        tickfont=dict(size=10, color="#6e6e78"),
    ),
)

event = st.plotly_chart(
    fig,
    use_container_width=True,
    on_select="rerun",
    selection_mode="points",
    key=f"cal_{st.session_state.cal_nonce}",
    config={"displayModeBar": False},
)

# ---- handle clicks: queue the dialog, reset the chart selection ----
cd = selected_customdata(event)
if cd:
    if cd[0] == "entry" and cd[1] is not None:
        st.session_state.pending = ("entry", cd[1])
        st.session_state.cal_nonce += 1
        st.rerun()
    elif cd[0] == "add" and can_edit:
        try:
            d = date.fromisoformat(str(cd[1]))
            h = int(cd[2])
            m = int(cd[3])
        except (TypeError, ValueError, IndexError):
            d, h, m = None, 9, 0
        if d:
            st.session_state.pending = ("add", d, max(0, min(23, h)),
                                        max(0, min(59, m)))
            st.session_state.cal_nonce += 1
            st.rerun()

# =====================================================
# No-start-time entries (from date-only sources like the MCP)
# =====================================================

if untimed:
    st.markdown("**No start time** <span style='color:#70757a;font-size:0.8rem;'>logged without a time - click to open</span>", unsafe_allow_html=True)
    u_cols = st.columns(min(len(untimed), 4))
    for j, e in enumerate(sorted(untimed, key=lambda x: x["dt"])):
        info = lookup.get(to_int(e.get("task_id")), LOOKUP_DEFAULT)
        lock = " (locked)" if is_true(e.get("approved")) else ""
        lbl = f"{e['dt'].strftime('%a %d')} - {fmt_dur(e['duration'])} - {info['client']}{lock}"
        if u_cols[j % len(u_cols)].button(lbl, key=f"u_{e['id']}", use_container_width=True):
            entry_dialog(e, lookup, editable=can_edit and not is_true(e.get("approved")))

if is_confirmed:
    st.caption("Week is validated: all entries are validated and read-only.")
elif wk_status == "submitted":
    st.caption("Week is submitted: the grid is read-only."
               + (" Use Unsubmit to make changes." if is_self else ""))
elif not can_edit:
    # Someone else's open week. An open week of your own needs no caption:
    # it is editable, which the grid already shows by being clickable.
    st.caption("Read-only view of this calendar.")